# coding=utf-8
import os
import sys
base_path = os.getcwd()
sys.path.append(base_path)
import openpyxl
from ExcelReader.handle_color import handle_color


class HandleExcel:
    '''
    获取表格路径
    '''

    def load_excel(self):
        open_excel = openpyxl.load_workbook('C:\\Users\\zhouk\\Desktop\\ued_class\\darkColor.xlsx')
        return open_excel

    def get_sheet_data(self, n):
        '''
        获取所有sheet中的内容
        '''
        sheet_name = self.load_excel().sheetnames
        excel_value = self.load_excel()[sheet_name[n]]
        return excel_value

    def get_cell_value(self, row, column, n):
        '''
        获取某一个单元格的内容
        '''
        cell_value = self.get_sheet_data(n).cell(row=row, column=column).value
        return cell_value

    def get_rows(self, n):
        '''
        获取行数
        '''
        row = self.get_sheet_data(n).max_row
        return row

    def get_row_data(self, row, n):
        '''
        获取某一行的数据
        '''
        row_list = []
        for i in self.get_sheet_data(n)[row]:
            row_list.append(i.value)
        return row_list

    def write_excel_data(self, row, clos, data, n):
        '''
        写入数据
        '''
        wb = self.load_excel()
        sheet_name = self.load_excel().sheetnames
        ws = wb[sheet_name[n]]
        ws.cell(row, clos, data)
        wb.save('D:\\dev\\python_project\\PythonApiTest\\Case\\YongLi.xlsx')

    def get_clos_data(self, n, key=None):
        '''
        获取某一列的数据
        '''
        clos_list = []
        if key == None:
            key = 'A'
        data = self.get_sheet_data(n)[key]
        for i in data:
            clos_list.append(i.value)

        return clos_list

    def get_row_number(self, case_id, n):
        '''
        获取某一行的行号
        '''
        clos_list = self.get_clos_data(n)
        num = 1
        for clos_data in clos_list:
            if clos_data == case_id:
                return num
            num = num + 1

        return num

    def get_excel_data(self, n):
        data_list = []
        for i in range(self.get_rows(n)):
            data = self.get_row_data((i + 2), n)
            data_list.append(data)
        return data_list


if __name__ == '__main__':
    handle_color = handle_color()
    handleexcel = HandleExcel()
    handle_color.build_html(handleexcel.get_excel_data(1))